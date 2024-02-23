import openpyxl
# from os.path import join,splitext
from os import listdir
import pdfplumber

def PDFToExcel(fileIn, fileOut):
    with pdfplumber.open(fileIn) as pdf:
        workbook = openpyxl.Workbook()
        sheet_index = 1
        page_count = 1

        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
            
        for page in pdf.pages:
            tables = page.find_tables()
            if tables:
                for table in tables:
                    print(f"Sheet index ... {sheet_index}/{len(pdf.pages)}")
                    xlsx_sheet = workbook.create_sheet(f"Sheet{sheet_index}")
                    header_row = table.extract()[0]
                    data_rows = table.extract()[1:]
                    for col_index, cell in enumerate(header_row):
                        xlsx_sheet.cell(row=1, column=col_index+1).value = cell
                    for row_index, row in enumerate(data_rows):
                        for col_index, cell in enumerate(row):
                            xlsx_sheet.cell(row=row_index+2, column=col_index+1).value = cell
                    sheet_index += 1
            page_count += 1

        workbook.save(fileOut)